"""
Excel Export Utility
Convert KK-OCR JSON output to Excel format
"""

import json
from typing import Dict, List, Optional, Union
from pathlib import Path
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("Warning: openpyxl not installed. Install with: pip install openpyxl")


def json_to_excel(
    json_data: Dict,
    output_path: str,
    include_metadata: bool = True
) -> str:
    """
    Convert single KK extraction result to Excel
    
    Args:
        json_data: Extraction result dictionary
        output_path: Output Excel file path
        include_metadata: Include processing metadata
        
    Returns:
        Path to created Excel file
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl is required for Excel export")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Data KK"
    
    # Styles
    header_font = Font(bold=True, size=14)
    subheader_font = Font(bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF")
    
    row = 1
    
    # Title
    ws.cell(row=row, column=1, value="DATA KARTU KELUARGA").font = header_font
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    row += 2
    
    # Header Data
    header = json_data.get("header", {})
    header_fields = [
        ("No. Kartu Keluarga", header.get("no_kk", "-")),
        ("Nama Kepala Keluarga", header.get("kepala_keluarga", "-")),
        ("Alamat", header.get("alamat", "-")),
        ("RT/RW", f"{header.get('rt', '-')}/{header.get('rw', '-')}"),
        ("Desa/Kelurahan", header.get("desa", header.get("desa_kelurahan", "-"))),
        ("Kecamatan", header.get("kecamatan", "-")),
        ("Kabupaten/Kota", header.get("kabupaten_kota", "-")),
        ("Provinsi", header.get("provinsi", "-")),
        ("Kode Pos", header.get("kode_pos", "-")),
        ("Tanggal Pembuatan", header.get("tanggal_pembuatan", "-")),
    ]
    
    for field_name, field_value in header_fields:
        ws.cell(row=row, column=1, value=field_name).font = Font(bold=True)
        ws.cell(row=row, column=2, value=field_value)
        row += 1
    
    row += 1
    
    # Family Members Table Header
    ws.cell(row=row, column=1, value="DAFTAR ANGGOTA KELUARGA").font = subheader_font
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=14)
    row += 1
    
    member_columns = [
        "No", "Nama Lengkap", "NIK", "Jenis Kelamin", "Tempat Lahir",
        "Tanggal Lahir", "Agama", "Pendidikan", "Pekerjaan",
        "Status Perkawinan", "Status Keluarga", "Kewarganegaraan",
        "Nama Ayah", "Nama Ibu"
    ]
    
    for col, header_text in enumerate(member_columns, 1):
        cell = ws.cell(row=row, column=col, value=header_text)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    row += 1
    
    # Family Members Data
    for i, member in enumerate(json_data.get("anggota_keluarga", []), 1):
        member_data = [
            i,
            member.get("nama_lengkap", "-"),
            member.get("nik", "-"),
            member.get("jenis_kelamin", "-"),
            member.get("tempat_lahir", "-"),
            member.get("tanggal_lahir", "-"),
            member.get("agama", "-"),
            member.get("pendidikan", "-"),
            member.get("jenis_pekerjaan", "-"),
            member.get("status_perkawinan", "-"),
            member.get("status_keluarga", "-"),
            member.get("kewarganegaraan", "-"),
            member.get("nama_ayah", "-"),
            member.get("nama_ibu", "-"),
        ]
        
        for col, value in enumerate(member_data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical='center')
        row += 1
    
    row += 2
    
    # Metadata (optional)
    if include_metadata:
        metadata = json_data.get("metadata", {})
        ws.cell(row=row, column=1, value="METADATA PEMROSESAN").font = subheader_font
        row += 1
        metadata_fields = [
            ("Timestamp", metadata.get("processing_timestamp", "-")),
            ("File Sumber", metadata.get("source_file", "-")),
            ("Model YOLO", metadata.get("model_version_yolo", "-")),
            ("Model UNet", metadata.get("model_version_unet", "-")),
            ("Model VLM", metadata.get("model_version_vlm", "-")),
        ]
        for field_name, field_value in metadata_fields:
            ws.cell(row=row, column=1, value=field_name)
            ws.cell(row=row, column=2, value=field_value)
            row += 1
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    wb.save(output_path)
    return output_path


def batch_to_excel(
    batch_response: Dict,
    output_path: str
) -> str:
    """
    Convert batch extraction results to Excel with multiple sheets
    
    Args:
        batch_response: Batch extraction response dictionary
        output_path: Output Excel file path
        
    Returns:
        Path to created Excel file
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl is required for Excel export")
    
    wb = Workbook()
    
    # Summary Sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    header_font = Font(bold=True, size=14)
    subheader_font = Font(bold=True, size=11)
    
    ws_summary.cell(row=1, column=1, value="BATCH PROCESSING SUMMARY").font = header_font
    ws_summary.merge_cells('A1:D1')
    
    summary = batch_response.get("summary", {})
    summary_data = [
        ("Job ID", batch_response.get("job_id", "-")),
        ("Status", batch_response.get("status", "-")),
        ("Total Files", summary.get("total_files", 0)),
        ("Successful", summary.get("successful", 0)),
        ("Failed", summary.get("failed", 0)),
        ("Total Time (seconds)", summary.get("total_time_seconds", 0)),
        ("Average Time/File", summary.get("average_time_per_file", 0)),
    ]
    
    row = 3
    for field, value in summary_data:
        ws_summary.cell(row=row, column=1, value=field).font = Font(bold=True)
        ws_summary.cell(row=row, column=2, value=value)
        row += 1
    
    row += 2
    
    # Results table
    ws_summary.cell(row=row, column=1, value="FILE RESULTS").font = subheader_font
    row += 1
    
    result_headers = ["No", "Filename", "Status", "Processing Time (s)", "Error"]
    for col, header in enumerate(result_headers, 1):
        ws_summary.cell(row=row, column=col, value=header).font = Font(bold=True)
    row += 1
    
    for i, result in enumerate(batch_response.get("results", []), 1):
        ws_summary.cell(row=row, column=1, value=i)
        ws_summary.cell(row=row, column=2, value=result.get("filename", "-"))
        ws_summary.cell(row=row, column=3, value=result.get("status", "-"))
        ws_summary.cell(row=row, column=4, value=result.get("processing_time_seconds", 0))
        ws_summary.cell(row=row, column=5, value=result.get("error", "-"))
        row += 1
    
    # Individual result sheets
    for i, result in enumerate(batch_response.get("results", [])):
        if result.get("status") == "success" and result.get("data"):
            # Sheet name max 31 chars
            sheet_name = f"{i+1}_{result['filename']}"[:31]
            ws = wb.create_sheet(title=sheet_name)
            
            # Add header data
            data = result["data"]
            header = data.get("header", {})
            
            ws.cell(row=1, column=1, value=f"KK: {header.get('no_kk', 'N/A')}").font = header_font
            
            row = 3
            for field, value in [
                ("Kepala Keluarga", header.get("kepala_keluarga", "-")),
                ("Alamat", header.get("alamat", "-")),
                ("Desa", header.get("desa", "-")),
            ]:
                ws.cell(row=row, column=1, value=field).font = Font(bold=True)
                ws.cell(row=row, column=2, value=value)
                row += 1
            
            row += 1
            
            # Family members
            ws.cell(row=row, column=1, value="Anggota Keluarga:").font = subheader_font
            row += 1
            
            member_headers = ["No", "Nama", "NIK", "Status Keluarga"]
            for col, h in enumerate(member_headers, 1):
                ws.cell(row=row, column=col, value=h).font = Font(bold=True)
            row += 1
            
            for j, member in enumerate(data.get("anggota_keluarga", []), 1):
                ws.cell(row=row, column=1, value=j)
                ws.cell(row=row, column=2, value=member.get("nama_lengkap", "-"))
                ws.cell(row=row, column=3, value=member.get("nik", "-"))
                ws.cell(row=row, column=4, value=member.get("status_keluarga", "-"))
                row += 1
    
    # Auto-adjust column widths for summary
    for col in ws_summary.columns:
        max_length = 0
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws_summary.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)
    
    wb.save(output_path)
    return output_path


def export_from_file(
    json_file: str,
    excel_file: Optional[str] = None,
    is_batch: bool = False
) -> str:
    """
    Export JSON file to Excel
    
    Args:
        json_file: Path to JSON file
        excel_file: Output Excel path (optional, auto-generated if not provided)
        is_batch: True if batch response, False for single extraction
        
    Returns:
        Path to created Excel file
    """
    with open(json_file, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    if excel_file is None:
        json_path = Path(json_file)
        excel_file = str(json_path.with_suffix('.xlsx'))
    
    if is_batch:
        return batch_to_excel(data, excel_file)
    else:
        return json_to_excel(data, excel_file)


if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description="Convert KK-OCR JSON to Excel")
    parser.add_argument("input", help="Input JSON file")
    parser.add_argument("-o", "--output", help="Output Excel file")
    parser.add_argument("--batch", action="store_true", help="Batch response format")
    
    args = parser.parse_args()
    
    output = export_from_file(args.input, args.output, args.batch)
    print(f"Excel file created: {output}")
